import io
import csv
import re
from typing import List, Dict, Any, Tuple
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from app.models import Worker, Department

class ImportService:
    @staticmethod
    def parse_file_rows(file_contents: bytes, filename: str) -> List[Dict[str, str]]:
        """Parse raw file bytes into list of dictionary row records (CSV or XLSX)."""
        rows: List[Dict[str, str]] = []
        if filename.endswith(".csv"):
            text_data = file_contents.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text_data))
            for row in reader:
                rows.append({k.strip(): (v.strip() if v else "") for k, v in row.items() if k})
        elif filename.endswith(".xlsx"):
            wb = load_workbook(filename=io.BytesIO(file_contents), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in ws[1] if cell.value]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_dict = {}
                    for i, h in enumerate(headers):
                        val = row[i] if i < len(row) else ""
                        row_dict[h] = str(val or "").strip()
                    rows.append(row_dict)
        else:
            raise ValueError("Unsupported file format. Please upload a .csv or .xlsx file.")
        return rows

    @staticmethod
    def validate_workers_import(db: Session, rows: List[Dict[str, str]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        Validate worker import rows.
        Returns (valid_records, warnings, errors).
        """
        valid_records: List[Dict[str, Any]] = []
        warnings: List[Dict[str, Any]] = []
        errors: List[Dict[str, Any]] = []

        existing_codes = {w.employee_number for w in db.query(Worker).all()}
        existing_emails = {w.email.lower() for w in db.query(Worker.email).filter(Worker.email.isnot(None)).all()}

        email_regex = re.compile(r"^[^@]+@[^@]+\.[^@]+$")

        for idx, row in enumerate(rows, start=2): # Row index starting at 2 (header = row 1)
            code = row.get("Employee Number") or row.get("Worker Code") or row.get("employee_number") or row.get("code") or row.get("ID")
            first_name = row.get("First Name") or row.get("first_name")
            last_name = row.get("Last Name") or row.get("last_name")
            email = row.get("Email") or row.get("email")
            dept_name = row.get("Department") or row.get("department")

            # Required fields check
            if not code or not first_name or not last_name:
                errors.append({
                    "row": idx,
                    "code": code or "UNKNOWN",
                    "error": "Missing required fields (Worker Code, First Name, or Last Name)."
                })
                continue

            # Duplicate Code check
            if code in existing_codes:
                warnings.append({
                    "row": idx,
                    "code": code,
                    "warning": f"Worker Code '{code}' already exists in database. Row will be skipped."
                })
                continue

            # Invalid / Duplicate Email check
            if email:
                if not email_regex.match(email):
                    errors.append({
                        "row": idx,
                        "code": code,
                        "error": f"Invalid email format '{email}'."
                    })
                    continue
                if email.lower() in existing_emails:
                    warnings.append({
                        "row": idx,
                        "code": code,
                        "warning": f"Email '{email}' is already associated with another worker. Row will be skipped."
                    })
                    continue

            valid_records.append({
                "row": idx,
                "code": code,
                "first_name": first_name,
                "last_name": last_name,
                "email": email or f"{code.lower()}@company.local",
                "dept_name": dept_name
            })

            # Add to set for duplicate tracking within the uploaded file itself
            existing_codes.add(code)
            if email:
                existing_emails.add(email.lower())

        return valid_records, warnings, errors

    @staticmethod
    def execute_workers_import(db: Session, valid_records: List[Dict[str, Any]]) -> int:
        """Transactionally insert validated worker records into PostgreSQL database."""
        inserted_count = 0
        departments = {d.name.lower(): d for d in db.query(Department).all()}
        default_dept = list(departments.values())[0] if departments else None

        try:
            for rec in valid_records:
                dept_obj = None
                if rec.get("dept_name"):
                    dept_obj = departments.get(rec["dept_name"].lower(), default_dept)
                else:
                    dept_obj = default_dept

                worker = Worker(
                    employee_number=rec["code"],
                    first_name=rec["first_name"],
                    last_name=rec["last_name"],
                    email=rec["email"],
                    department_id=dept_obj.id if dept_obj else None,
                    active=True,
                    weekly_contract_hours=40,
                )
                db.add(worker)
                inserted_count += 1
            
            db.commit()
            return inserted_count
        except Exception as e:
            db.rollback()
            raise e
