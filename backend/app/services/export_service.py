import io
import csv
import json
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Schedule, Assignment, Worker, ShiftInstance, ShiftType, AuditLog


class ExportService:
    """Service responsible for generating CSV, Excel, and JSON exports of schedule data."""

    @staticmethod
    def _schedule_label(schedule: Schedule) -> str:
        """Return a human-readable label for a schedule period."""
        import calendar
        month_name = calendar.month_name[schedule.month]
        return f"{month_name} {schedule.year}"

    @staticmethod
    def export_schedule_csv(db: Session, schedule: Schedule) -> str:
        """Export complete schedule assignments to a UTF-8 CSV string."""
        output = io.StringIO()
        writer = csv.writer(output)

        label = ExportService._schedule_label(schedule)

        # Header block
        writer.writerow(["Schedule ID", schedule.id])
        writer.writerow(["Period", label])
        writer.writerow(["Status", schedule.status.value if hasattr(schedule.status, "value") else schedule.status])
        writer.writerow([])

        # Column headers
        writer.writerow([
            "Date", "Shift Name", "Start Time", "End Time",
            "Employee Number", "First Name", "Last Name", "Locked", "Notes"
        ])

        for instance in sorted(schedule.shift_instances, key=lambda i: i.date):
            for assignment in instance.assignments:
                worker = assignment.worker
                shift_type = instance.shift_type
                writer.writerow([
                    str(instance.date),
                    shift_type.name if shift_type else "",
                    str(shift_type.start_time) if shift_type else "",
                    str(shift_type.end_time) if shift_type else "",
                    worker.employee_number if worker else "",
                    worker.first_name if worker else "",
                    worker.last_name if worker else "",
                    "YES" if assignment.locked else "NO",
                    assignment.notes or "",
                ])

        return output.getvalue()

    @staticmethod
    def export_schedule_json(db: Session, schedule: Schedule) -> Dict[str, Any]:
        """Export complete schedule to a machine-readable JSON structure."""
        assignments = []
        for instance in schedule.shift_instances:
            for a in instance.assignments:
                assignments.append({
                    "assignment_id": str(a.id),
                    "date": str(instance.date),
                    "shift_name": instance.shift_type.name if instance.shift_type else None,
                    "worker_id": str(a.worker_id),
                    "employee_number": a.worker.employee_number if a.worker else None,
                    "first_name": a.worker.first_name if a.worker else None,
                    "last_name": a.worker.last_name if a.worker else None,
                    "locked": a.locked,
                })

        return {
            "schedule": {
                "id": str(schedule.id),
                "month": schedule.month,
                "year": schedule.year,
                "period": ExportService._schedule_label(schedule),
                "status": schedule.status.value if hasattr(schedule.status, "value") else str(schedule.status),
            },
            "assignments_count": len(assignments),
            "assignments": assignments,
        }

    @staticmethod
    def export_schedule_excel(db: Session, schedule: Schedule) -> bytes:
        """Export complete schedule to a formatted Excel workbook (.xlsx)."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule Grid"

        # Styling
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        title_font = Font(name="Calibri", size=14, bold=True)
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        label = ExportService._schedule_label(schedule)

        # Title block
        ws.append(["Staff Scheduler — Official Schedule Export"])
        ws.cell(row=1, column=1).font = title_font
        ws.append(["Period:", label])
        ws.append(["Status:", schedule.status.value if hasattr(schedule.status, "value") else str(schedule.status)])
        ws.append([])

        # Column headers
        headers = [
            "Date", "Shift Name", "Start Time", "End Time",
            "Employee Number", "First Name", "Last Name", "Locked", "Notes"
        ]
        ws.append(headers)
        header_row_idx = 5

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for instance in sorted(schedule.shift_instances, key=lambda i: i.date):
            for assignment in instance.assignments:
                worker = assignment.worker
                shift_type = instance.shift_type
                ws.append([
                    str(instance.date),
                    shift_type.name if shift_type else "",
                    str(shift_type.start_time) if shift_type else "",
                    str(shift_type.end_time) if shift_type else "",
                    worker.employee_number if worker else "",
                    worker.first_name if worker else "",
                    worker.last_name if worker else "",
                    "YES" if assignment.locked else "NO",
                    assignment.notes or "",
                ])

        # Apply borders and auto-width
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_workers_csv(db: Session) -> str:
        """Export worker roster to CSV using correct model field names."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Employee Number", "First Name", "Last Name", "Email",
            "Department", "Active", "Contract Hours/Week"
        ])
        workers = db.query(Worker).order_by(Worker.employee_number).all()
        for w in workers:
            writer.writerow([
                w.employee_number,
                w.first_name,
                w.last_name,
                w.email or "",
                w.department.name if w.department else "",
                "YES" if w.active else "NO",
                w.weekly_contract_hours or 40,
            ])
        return output.getvalue()

    @staticmethod
    def export_audit_log_csv(db: Session) -> str:
        """Export system audit log entries to CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Timestamp", "User", "Action", "Entity Type", "Entity ID", "Details"])
        logs = db.query(AuditLog).order_by(AuditLog.created_at.desc()).limit(1000).all()
        for log in logs:
            writer.writerow([
                str(log.created_at),
                log.user.username if log.user else "System",
                log.action,
                log.entity_type,
                log.entity_id or "",
                log.reason or "",
            ])
        return output.getvalue()
